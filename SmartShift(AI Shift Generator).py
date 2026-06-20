# =====================================================
# IMPORTS, INITIALIZATION & USER INPUT MODULE
# =====================================================
import os, time, numpy as np, pandas as pd
from datetime import datetime, timedelta
from sklearn.metrics import r2_score, root_mean_squared_error
from xgboost import XGBRegressor
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
print("=======================================================================\n🪐 ENOC SMARTSHIFT: INTERACTIVE PRODUCTION LABOUR PIPELINE\n=======================================================================")
print("🏢 Real-World Infrastructure Context:\n• Active UAE ENOC Network Footprint : 210+ Service Stations\n• Network Strategic Growth Target   : 250 Service Stations\n• Managed Cluster Sample Volume     : 4 High-Density Flagship Sites\n=======================================================================\n")
print("⌨️ [USER INTERFACE]: Enter Operational Parameters\n" + "-" * 50 + "\n🎪 EVENT CONFIGURATION\n" + "-" * 50)
while True:
    event_exists = input("Is there any special event during this scheduling period? (Y/N): ").strip().upper()
    if event_exists in ["Y", "N"]: break
    print("❌ Please enter Y or N.")
event_name, event_date = "No Event", None
if event_exists == "Y":
    while True:
        event_name = input("Enter Event Name: ").strip()
        if len(event_name) > 0: break
        print("❌ Event name cannot be empty.")
    while True:
        try: event_date = datetime.strptime(input("Enter Event Date (YYYY-MM-DD): ").strip(), "%Y-%m-%d").date(); break
        except ValueError: print("❌ Invalid format. Use YYYY-MM-DD.")
while True:
    try: start_date = datetime.strptime(input("📅 Enter the Schedule Start Date (YYYY-MM-DD, e.g., 2026-06-10): ").strip(), "%Y-%m-%d"); print("✓ Start date format validated successfully."); break  
    except ValueError: print("❌ INVALID FORMAT: The system requires the exact YYYY-MM-DD format (e.g., 2026-08-25).\n   Please try entering the scheduling horizon date again.\n")
valid_sites = ['SITE-1044', 'SITE-1102', 'SITE-1089', 'SITE-1215']
while True:
    user_site = input("⛽ Enter Station ID (SITE-1044, SITE-1102, SITE-1089, SITE-1215): ").strip().upper()
    if user_site in valid_sites: print(f"✓ Station {user_site} validated against registered asset registry."); break  
    else: print(f"❌ ERROR: '{user_site}' is not under the available ENOC sites.\n   Please enter one of the verified training cluster sites: {valid_sites}\n")
print("\n" + "="*70 + "\n")
# =====================================================
# CONFIGURATION & DATA SOURCE VALIDATION
# =====================================================
site_profiles = {'SITE-1044': 'Near_Major_Mall', 'SITE-1102': 'Highway_Transit_Hub', 'SITE-1089': 'Residential_Community', 'SITE-1215': 'Commercial_Business_Hub'}
csv_data_path, calendar_path = r"C:\Portfolio-ML\Datasets\enoc_realistic_dataset.csv", r"C:\Portfolio-ML\Datasets\uae_calendar_2026.csv"
if not os.path.exists(csv_data_path): print(f"⚠️ ERROR: Dataset not found in specified file path: {csv_data_path}"); exit()
if not os.path.exists(calendar_path): print(f"⚠️ ERROR: UAE Holiday Corporate Calendar not found at: {calendar_path}"); exit()
# =====================================================
# DATA LOADING, FEATURE ENGINEERING & TRAIN-TEST SPLIT
# =====================================================
print("Phase: Loading the Dataset..."); start_time = time.time()
df = pd.read_csv(csv_data_path); print(f"✓ Dataset loaded into memory in {time.time() - start_time:.2f} seconds.\nProcessing date-time properties and labels...")
df["timestamp"] = pd.to_datetime(df["timestamp"])
df["year"], df["month"], df["hour_of_day"], df["day_of_week"] = df["timestamp"].dt.year, df["timestamp"].dt.month, df["timestamp"].dt.hour, df["timestamp"].dt.dayofweek
df["day_of_month"], df["week_of_year"], df["quarter"] = df["timestamp"].dt.day, df["timestamp"].dt.isocalendar().week.astype(int), df["timestamp"].dt.quarter
df = df.sort_values(["station_id", "timestamp"])
Y, Y1 = df["required_staff_count"], df["historical_average_customer_count"]
X, X1 = df.drop(columns=["required_staff_count", "timestamp"]), df.drop(columns=["historical_average_customer_count", "required_staff_count", "timestamp"])
for col in ["station_id", "location_profile"]:
    if col in X.columns: X[col] = X[col].astype("category")
    if col in X1.columns: X1[col] = X1[col].astype("category")
X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.20, random_state=42)
X1_train, X1_test, Y1_train, Y1_test = train_test_split(X1, Y1, test_size=0.20, random_state=42)
# =====================================================
# MACHINE LEARNING ENGINE & EVALUATION
# =====================================================
print("Initializing and fitting the XGBoost regression engine..."); train_start = time.time()
m, m1 = XGBRegressor(enable_categorical=True, random_state=42, n_jobs=-1), XGBRegressor(enable_categorical=True, random_state=42, n_jobs=-1)
m.fit(X_train, Y_train); m1.fit(X1_train, Y1_train)
print(f"✓ Model training complete in {time.time() - train_start:.2f} seconds.")
Y_pred, Y1_pred = m.predict(X_test), m1.predict(X1_test)
r2, r21 = r2_score(Y_test, Y_pred), r2_score(Y1_test, Y1_pred)
rmse, rmse1 = root_mean_squared_error(Y_test, Y_pred), root_mean_squared_error(Y1_test, Y1_pred)
print(f"📈 Background Engine Verification Metrics(Model 1): R2 = {r2*100:.2f}% | RMSE = {rmse:.2f} \n📈 Background Engine Verification Metrics(Model 2): R2 = {r21*100:.2f}% | RMSE = {rmse1:.2f} \n\n" + "="*70 + "\n")
# =====================================================
# FUTURE FORECAST GENERATION & MAIN ENGINE
# =====================================================
forecast_records, calendar_lookup = [], pd.read_csv(calendar_path)
for h in range(168):
    current_time = start_date + timedelta(hours=h); current_date = current_time.date()
    date_str, hour_num, day_num = current_time.strftime("%Y-%m-%d"), current_time.hour, current_time.weekday()
    day_match = calendar_lookup[calendar_lookup["date_string"] == date_str]
    is_holiday, holiday_label = (int(day_match["is_public_holiday"].values[0]), str(day_match["holiday_name"].values[0])) if not day_match.empty else (0, "Normal Day")
    is_weekend, is_event = int(day_num in [4, 5, 6]), int(event_exists == "Y" and current_time.date() == event_date)
    forecast_records.append({
        "Timestamp_String": current_time.strftime("%Y-%m-%d %H:%M:%S"), "month": current_time.month, "station_id": user_site, "location_profile": site_profiles[user_site],
        "is_weekend": is_weekend, "is_public_holiday": is_holiday, "is_event_nearby": is_event, "Event_Name": event_name if is_event else "No Event",
        "historical_average_customer_count": 0, "Holiday_Name": holiday_label, "day_of_month": current_time.day, "week_of_year": int(current_time.isocalendar().week),
        "quarter": ((current_time.month - 1) // 3) + 1, "year": current_time.year, "hour_of_day": hour_num, "day_of_week": day_num
    })
X_future = pd.DataFrame(forecast_records)
customer_features = X_future.drop(columns=["Timestamp_String", "Holiday_Name", "Event_Name", "historical_average_customer_count"])
customer_features["station_id"], customer_features["location_profile"] = customer_features["station_id"].astype("category"), customer_features["location_profile"].astype("category")
X_future["historical_average_customer_count"] = m1.predict(customer_features[X1.columns])
staff_features = X_future.drop(columns=["Timestamp_String", "Holiday_Name", "Event_Name"])
staff_features["station_id"], staff_features["location_profile"] = staff_features["station_id"].astype("category"), staff_features["location_profile"].astype("category")
X_future["Predicted_Staff_Needed"] = np.clip(np.round(m.predict(staff_features[X.columns])), 2, 16).astype(int)
X_future["Date"] = pd.to_datetime(X_future["Timestamp_String"]).dt.date
daily_mean, daily_std = X_future.groupby("Date")["historical_average_customer_count"].transform("mean"), X_future.groupby("Date")["historical_average_customer_count"].transform("std")
X_future["Is_Rush"] = (X_future["historical_average_customer_count"] > (daily_mean + 1.75 * daily_std))
# =====================================================
# EMPLOYEE MASTER & ROSTER GENERATION (OUTPUT 1)
# =====================================================
employees = pd.DataFrame({"Employee_ID": [f"EMP-{i}" for i in range(1001, 1040)], "Employee_Type": ["Permanent"] * 30 + ["Floating"] * 9, "Home_Site": [user_site] * 39})
employees["Role"] = ["Supervisor"] * 3 + ["Cashier"] * 8 + ["Fuel_Attendant"] * 20 + ["Store_Assistant"] * 8
print("\n📅 [OUTPUT 1]: Mapping predictions into final employee shift rosters...\n" + "-" * 85)
timetable_records, permanent_capacity = [], 12
for idx, row in X_future.iterrows():
    needed = int(row["Predicted_Staff_Needed"])
    hour_num = int(row["hour_of_day"])
    timetable_records.append({
        "Date": row["Timestamp_String"][:10], "Hour_Block": f"{hour_num:02d}:00 - {(hour_num + 1) % 24:02d}:00", "Forecasted_Customers": round(row["historical_average_customer_count"]),
        "Permanent_Staff": min(needed, permanent_capacity), "Floating_Staff": max(0, needed - permanent_capacity), "Total_Staff_Required": needed, "Peak_Hour": "⚠️ PEAK RUSH" if row["Is_Rush"] else "NO"
    })
hourly_forecast_df = pd.DataFrame(timetable_records)
# =====================================================
# EXCEL EXPORT & SHEET FORMATTING ENGINE
# =====================================================
total_labour_cost = X_future["Predicted_Staff_Needed"].sum() * 30
excel_output_path = r'c:\Portfolio-ML\Internship\enoc_weekly_workforce_planning.xlsx'
peak_staff = int(X_future["Predicted_Staff_Needed"].max())
floating_staff_required = max(0, peak_staff - permanent_capacity)
summary_df = pd.DataFrame({
    "Metric": ["Forecasted Customers", "Peak Staff Requirement", "Average Staff Requirement", "Rush Hours Detected", "Floating Staff Required", "Estimated Labour Cost (AED)", "Permanent Staff Capacity", "Peak Utilization (%)"],
    "Value": [int(X_future["historical_average_customer_count"].sum()), peak_staff, round(X_future["Predicted_Staff_Needed"].mean(), 2), int(X_future["Is_Rush"].sum()), floating_staff_required, round(total_labour_cost, 2), permanent_capacity, round((peak_staff / permanent_capacity) * 100, 1)]
})
with pd.ExcelWriter(excel_output_path, engine="openpyxl") as writer:
    summary_df.to_excel(writer, sheet_name="Executive Summary", index=False)
    hourly_forecast_df.to_excel(writer, sheet_name="Hourly Forecast", index=False)
wb = load_workbook(excel_output_path)
yellow_fill, header_fill = PatternFill(fill_type="solid", fgColor="FFF2CC"), PatternFill(fill_type="solid", fgColor="007A87")
thin_side, thick_side = Side(border_style="thin", color="CCCCCC"), Side(border_style="medium", color="007A87")
header_border, regular_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side), Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
header_font, regular_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF"), Font(name="Segoe UI", size=10, color="333333")
for ws in wb.worksheets:
    for cell in ws[1]: cell.fill, cell.font, cell.border = header_fill, header_font, header_border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        peak_row = ws.title == "Hourly Forecast" and str(row[-1].value) in ["⚠️ PEAK RUSH", "YES"]
        for cell in row:
            cell.font, cell.border, cell.alignment = regular_font, regular_border, Alignment(horizontal="center", vertical="center")
            if peak_row: cell.fill = yellow_fill
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 18)
hourly_ws = wb["Hourly Forecast"]
for row in range(2, hourly_ws.max_row):
    if hourly_ws.cell(row=row, column=1).value != hourly_ws.cell(row=row + 1, column=1).value:
        for col in range(1, hourly_ws.max_column + 1):
            cell = hourly_ws.cell(row=row, column=col)
            cell.border = Border(left=cell.border.left, right=cell.border.right, top=cell.border.top, bottom=thick_side)
wb.save(excel_output_path)
print(f"👉 Success! Excel file saved successfully.\nDestination Path: {excel_output_path}")
# =====================================================
# ANALYSIS OUTPUTS (OUTPUT 2, 4, 5)
# =====================================================
total_workforce_hours, total_employees = X_future["Predicted_Staff_Needed"].sum(), len(employees)
avg_hours_per_employee = round(total_workforce_hours / total_employees, 1)
print(f"\n🛡️ [OUTPUT 2]: Workforce Utilization & Overtime Risk Analysis...\n" + "-" * 70)
print(f"Weekly Workforce Hours Required : {total_workforce_hours}\nTotal Employees Available : {total_employees}\nAverage Employee Hours : {avg_hours_per_employee}")
print(f"\n⚠ Labour Overtime Risk Detected\nOvertime : {max(0, round(avg_hours_per_employee - 40, 1))} hours/employee" if avg_hours_per_employee > 40 else "\n✓ No Labour Overtime Risk Detected")
print(f"\nPermanent Capacity : {permanent_capacity}\nPeak Requirement : {peak_staff}\nFloating Staff Needed : {max(0, peak_staff - permanent_capacity)}")
print(f"\n🏆 [OUTPUT 5]: Post-Shift Efficiency & Statistical Accuracy Scorecard...\n📈 LIVE MODEL EVALUATION: System deployed successfully. Final R2 Score: {r2*100:.2f}%")
